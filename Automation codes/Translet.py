import openpyxl
from googletrans import Translator

# Load the Excel workbook
file_path = r"C:\Users\Sreedharan BR\Documents\Exce_folder.xlsx"  # Replace with your Excel file path
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Initialize the translator
translator = Translator()

# Define target languages (e.g., French, German, Spanish)
languages = ['mr', 'bn']  # ISO 639-1 language codes

# Translate content
for row in sheet.iter_rows(min_row=2, max_col=2):  # Assuming data is in the first column
    for cell in row:
        if cell.value:  # Check if cell is not empty
            translations = {}
            for lang in languages:
                try:
                    translated_text = translator.translate(cell.value, dest=lang).text
                    translations[lang] = translated_text
                except Exception as e:
                    translations[lang] = f"Error: {e}"
            
            # Add translations to the adjacent columns
            col_offset = 3  # Starting from the second column
            for i, lang in enumerate(languages):
                sheet.cell(row=cell.row, column=col_offset + i, value=translations[lang])

# Save the translated Excel file
output_file_path = r"C:\Users\Sreedharan BR\Documents\Book2.xlsx"
wb.save(output_file_path)

print(f"Translations saved to {output_file_path}")
